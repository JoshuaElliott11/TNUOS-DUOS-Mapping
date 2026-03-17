import argparse
import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from pyproj import Transformer

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except Exception:  # pragma: no cover
    load_workbook = None
    Workbook = None

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

DNO_FILE = DATA_DIR / "dno_20240503.geojson"
GSP_FILE = DATA_DIR / "gsp_20251204.geojson"
GEN_FILE = DATA_DIR / "tnuos_generation_zones.geojson"

Point = Tuple[float, float]
Ring = List[Point]
Polygon = List[Ring]
MultiPolygon = List[Polygon]


def _iter_polygons(geometry: Dict[str, Any]) -> Iterable[Polygon]:
    gtype = geometry.get("type")
    coords = geometry.get("coordinates")
    if not coords:
        return
    if gtype == "Polygon":
        yield coords
    elif gtype == "MultiPolygon":
        for polygon in coords:
            yield polygon


def _ring_bbox(ring: Ring) -> Tuple[float, float, float, float]:
    xs = [p[0] for p in ring]
    ys = [p[1] for p in ring]
    return min(xs), min(ys), max(xs), max(ys)


def _point_in_ring(point: Point, ring: Ring) -> bool:
    x, y = point
    inside = False
    n = len(ring)
    if n < 3:
        return False

    for i in range(n):
        x1, y1 = ring[i]
        x2, y2 = ring[(i + 1) % n]

        # Ray-casting toggle when edge crosses horizontal line at y.
        intersects = ((y1 > y) != (y2 > y))
        if intersects:
            xinters = (x2 - x1) * (y - y1) / ((y2 - y1) or 1e-20) + x1
            if x < xinters:
                inside = not inside

    return inside


def _point_in_polygon(point: Point, polygon: Polygon) -> bool:
    if not polygon:
        return False

    outer = polygon[0]
    if not _point_in_ring(point, outer):
        return False

    # If point falls inside a hole, it's outside the polygon.
    for hole in polygon[1:]:
        if _point_in_ring(point, hole):
            return False

    return True


def _load_geojson_features(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    features: List[Dict[str, Any]] = []
    for feature in data.get("features", []):
        geometry = feature.get("geometry") or {}
        if not geometry.get("coordinates"):
            continue

        polygons = list(_iter_polygons(geometry))
        if not polygons:
            continue

        bboxes = [_ring_bbox(poly[0]) for poly in polygons if poly and poly[0]]
        if not bboxes:
            continue

        minx = min(b[0] for b in bboxes)
        miny = min(b[1] for b in bboxes)
        maxx = max(b[2] for b in bboxes)
        maxy = max(b[3] for b in bboxes)

        features.append(
            {
                "properties": feature.get("properties", {}),
                "polygons": polygons,
                "bbox": (minx, miny, maxx, maxy),
            }
        )

    return features


def _contains_feature(point: Point, item: Dict[str, Any]) -> bool:
    x, y = point
    minx, miny, maxx, maxy = item["bbox"]
    if not (minx <= x <= maxx and miny <= y <= maxy):
        return False

    for polygon in item["polygons"]:
        if _point_in_polygon(point, polygon):
            return True
    return False


class RegionLookup:
    def __init__(self) -> None:
        self.dno_features = _load_geojson_features(DNO_FILE)
        self.gsp_features = _load_geojson_features(GSP_FILE)
        self.gen_features = _load_geojson_features(GEN_FILE)
        self.wgs84_to_bng = Transformer.from_crs("EPSG:4326", "EPSG:27700", always_xy=True)

    def _find(self, point: Point, features: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        for item in features:
            if _contains_feature(point, item):
                return item["properties"]
        return None

    def lookup(self, latitude: float, longitude: float) -> Dict[str, Any]:
        point_wgs84 = (longitude, latitude)
        dno_point = self.wgs84_to_bng.transform(longitude, latitude)

        dno = self._find(dno_point, self.dno_features) or {}
        gsp = self._find(point_wgs84, self.gsp_features) or {}
        gen = self._find(point_wgs84, self.gen_features) or {}

        return {
            "latitude": latitude,
            "longitude": longitude,
            "duos_region": dno.get("Area"),
            "duos_operator": dno.get("DNO_Full") or dno.get("DNO"),
            "tnuos_demand_region": gsp.get("GSPGroup"),
            "gsp_code": gsp.get("GSPs"),
            "tnuos_generation_zone": gen.get("layer"),
        }


def _find_lat_lon_keys(columns: Sequence[str]) -> Tuple[str, str]:
    normalized = {c.strip().lower(): c for c in columns}

    lat_candidates = ["latitude", "lat", "y"]
    lon_candidates = ["longitude", "lon", "lng", "long", "x"]

    lat_key = next((normalized[c] for c in lat_candidates if c in normalized), None)
    lon_key = next((normalized[c] for c in lon_candidates if c in normalized), None)

    if not lat_key or not lon_key:
        raise ValueError(
            "Could not find latitude/longitude columns. "
            "Expected one of latitude/lat and longitude/lon/lng/long."
        )

    return lat_key, lon_key


def process_csv(input_path: Path, output_path: Path, lookup: RegionLookup) -> None:
    with input_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row.")

    lat_key, lon_key = _find_lat_lon_keys(reader.fieldnames)

    out_rows: List[Dict[str, Any]] = []
    for row in rows:
        lat = float(row[lat_key])
        lon = float(row[lon_key])
        result = lookup.lookup(lat, lon)

        row.update(
            {
                "duos_region": result["duos_region"],
                "duos_operator": result["duos_operator"],
                "tnuos_demand_region": result["tnuos_demand_region"],
                "gsp_code": result["gsp_code"],
                "tnuos_generation_zone": result["tnuos_generation_zone"],
            }
        )
        out_rows.append(row)

    fieldnames = list(out_rows[0].keys()) if out_rows else list(reader.fieldnames)
    for col in [
        "duos_region",
        "duos_operator",
        "tnuos_demand_region",
        "gsp_code",
        "tnuos_generation_zone",
    ]:
        if col not in fieldnames:
            fieldnames.append(col)

    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)


def process_xlsx(input_path: Path, output_path: Path, lookup: RegionLookup) -> None:
    if load_workbook is None or Workbook is None:
        raise RuntimeError("openpyxl is required for Excel support. Install with: pip install openpyxl")

    wb = load_workbook(input_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if not header:
        raise ValueError("Excel sheet has no header row.")

    str_header = [str(h).strip() if h is not None else "" for h in header]
    lat_key, lon_key = _find_lat_lon_keys(str_header)
    lat_idx = str_header.index(lat_key) + 1
    lon_idx = str_header.index(lon_key) + 1

    out_cols = [
        "duos_region",
        "duos_operator",
        "tnuos_demand_region",
        "gsp_code",
        "tnuos_generation_zone",
    ]

    col_index: Dict[str, int] = {}
    max_col = ws.max_column
    for col in out_cols:
        if col in str_header:
            col_index[col] = str_header.index(col) + 1
        else:
            max_col += 1
            ws.cell(row=1, column=max_col, value=col)
            col_index[col] = max_col

    for row_num in range(2, ws.max_row + 1):
        lat_cell = ws.cell(row=row_num, column=lat_idx).value
        lon_cell = ws.cell(row=row_num, column=lon_idx).value
        if lat_cell is None or lon_cell is None:
            continue

        lat = float(lat_cell)
        lon = float(lon_cell)
        result = lookup.lookup(lat, lon)

        ws.cell(row=row_num, column=col_index["duos_region"], value=result["duos_region"])
        ws.cell(row=row_num, column=col_index["duos_operator"], value=result["duos_operator"])
        ws.cell(row=row_num, column=col_index["tnuos_demand_region"], value=result["tnuos_demand_region"])
        ws.cell(row=row_num, column=col_index["gsp_code"], value=result["gsp_code"])
        ws.cell(row=row_num, column=col_index["tnuos_generation_zone"], value=result["tnuos_generation_zone"])

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Lookup DUoS and TNUoS regions from latitude/longitude using NESO GIS boundaries."
    )
    parser.add_argument("--lat", type=float, help="Latitude in decimal degrees")
    parser.add_argument("--lon", type=float, help="Longitude in decimal degrees")
    parser.add_argument("--input", type=Path, help="Input CSV or XLSX for batch processing")
    parser.add_argument("--output", type=Path, help="Output CSV or XLSX path for batch processing")

    args = parser.parse_args()

    lookup = RegionLookup()

    if args.lat is not None and args.lon is not None:
        print(json.dumps(lookup.lookup(args.lat, args.lon), indent=2))
        return

    if args.input and args.output:
        suffix = args.input.suffix.lower()
        if suffix == ".csv":
            process_csv(args.input, args.output, lookup)
        elif suffix in {".xlsx", ".xlsm"}:
            process_xlsx(args.input, args.output, lookup)
        else:
            raise ValueError("Input file must be .csv or .xlsx/.xlsm")

        print(f"Wrote results to: {args.output}")
        return

    raise ValueError("Use either --lat/--lon for single lookup, or --input/--output for batch mode.")


if __name__ == "__main__":
    main()
